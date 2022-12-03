from openpyxl import load_workbook, styles
from contextlib import closing
from os.path import abspath
import os


class ResultWriter:

    def __init__(self, file_name):
        cur_dir = os.path.dirname(os.path.realpath(__file__))
        self.filename = cur_dir + '\\' + file_name
        self.wb = load_workbook(self.filename)
        self.ws = self.wb['Execute']

    def get_coordinates(self, text):
        """
        Search for the text in column A and return coordinates resulting column D
        """
        for cell in self.ws['A']:
            if cell.value is not None:
                if text in cell.value:
                    cell_cords = "D{}".format(cell.row)
        return cell_cords

    def add_value(self, cell_cords, value, report_path):
        try:
            with closing(load_workbook(filename=self.filename)) as wb:
                ws = wb['Execute']
                ws[cell_cords] = value
                ws[cell_cords.replace("D", "E")] = '=HYPERLINK("{}", "{}")'.format(report_path, ""
                                                                                                "Link")
                wb.save(self.filename)
        except:
            self.closeFile()
            with closing(load_workbook(filename=self.filename)) as wb:
                ws = wb['Execute']
                ws[cell_cords] = value
                ws[cell_cords.replace("D", "E")] = '=HYPERLINK("{}", "{}")'.format(report_path, "Report Link")
                wb.save(self.filename)

    def closeFile(self):
        try:
            os.system('TASKKILL /F /IM excel.exe')
        except:
            print("Unable to close {}".format(self.filename))
