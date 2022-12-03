import pandas as pd
import os
from openpyxl import load_workbook
from functools import reduce


class TestReader:
    def __init__(self,file):
        self.data_dict = {}
        cur_dir = os.path.dirname(os.path.realpath(__file__))
        self.file = cur_dir + '\\' + file
        self.df = pd.read_excel(self.file, sheet_name='Execute', engine='openpyxl')
        self.wb = load_workbook(filename=self.file, read_only=True)
        self.ws = self.wb['Input']

    def _get_enabled_tests(self):
        """
        This method will return the enabled/selected testcases from "Execute" sheet.
        """
        n_row = self.df.values.tolist()
        test_to_run = list(filter(None, map(lambda x: x if True in x else [], n_row)))
        required_data = []
        for test in test_to_run:
            required_data.append(test[0:2])
        return required_data

    def _get_test_param(self, start_row, end_row):
        data_rows = []
        for row in self.ws[start_row:end_row]:
            data_cols = []
            for cell in row:
                data_cols.append(cell.value)
            data_rows.append(data_cols)
        #print(data_rows)
        #data_rows = reduce(lambda x, y: x + y, data_rows)
        #print(data_rows)
        return data_rows

    def testcase_formation(self):
        en_tests = self._get_enabled_tests()
        test_name = [item[0] for item in en_tests]
        en_val = [item[1] for item in en_tests]
        for test in range(len(test_name)):
            testcase_name = test_name[test]
            par1 = en_val[test].split(",")[0]
            par2 = en_val[test].split(",")[1]
            test_list = self._get_test_param(par1, par2)
            self.data_dict[testcase_name] = test_list
        return self.data_dict
